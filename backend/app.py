"""Flask API for the free concert ticket tracker."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from backend.db import get_conn, init_db
from backend.kpi import calculate_portfolio_stats
from backend.scraper import scrape_tickpick_concert
from backend.signals import generate_buy_signal

app = Flask(__name__)
CORS(app)


@app.get("/api/health")
def health() -> tuple[dict, int]:
    return {"status": "ok"}, 200


@app.post("/api/concerts")
def add_concert():
    data = request.get_json(force=True)
    artist = data["artist"].strip()
    venue = data["venue"].strip()
    event_date = data["event_date"]
    city = data.get("city")

    concert_id = f"tickpick_{artist}_{venue}_{event_date}".lower().replace(" ", "_")

    with get_conn() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO concerts (id, name, artist, venue, city, event_date, tracking_start, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'active')
            """,
            (
                concert_id,
                f"{artist} @ {venue}",
                artist,
                venue,
                city,
                event_date,
                datetime.now(timezone.utc).isoformat(),
            ),
        )

    snapshot = scrape_tickpick_concert(artist, venue, event_date)
    _insert_snapshot(concert_id, snapshot)

    return jsonify({"status": "success", "concert_id": concert_id})


@app.get("/api/concerts")
def list_concerts():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM concerts WHERE status='active' ORDER BY event_date ASC").fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/concerts/<concert_id>/refresh")
def refresh_concert(concert_id: str):
    with get_conn() as conn:
        concert = conn.execute("SELECT * FROM concerts WHERE id = ?", (concert_id,)).fetchone()
    if not concert:
        return jsonify({"status": "error", "message": "Concert not found"}), 404

    snapshot = scrape_tickpick_concert(concert["artist"], concert["venue"], concert["event_date"])
    _insert_snapshot(concert_id, snapshot)
    return jsonify({"status": "success", "concert_id": concert_id, "snapshot": snapshot})


@app.get("/api/concerts/<concert_id>/price-history")
def price_history(concert_id: str):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT timestamp, get_in_price, median_price, listings_count, price_change_24h_pct, price_change_7d_pct
            FROM price_snapshots
            WHERE concert_id = ?
            ORDER BY timestamp ASC
            """,
            (concert_id,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/dashboard")
def dashboard():
    with get_conn() as conn:
        concerts = conn.execute(
            """
            SELECT c.*, p.get_in_price, p.median_price, p.listings_count, p.timestamp AS last_snapshot_at
            FROM concerts c
            LEFT JOIN price_snapshots p ON p.id = (
                SELECT ps.id
                FROM price_snapshots ps
                WHERE ps.concert_id = c.id
                ORDER BY ps.timestamp DESC
                LIMIT 1
            )
            WHERE c.status='active'
            ORDER BY c.event_date ASC
            """
        ).fetchall()
    return jsonify([dict(r) for r in concerts])


@app.get("/api/concerts/<concert_id>/signal")
def signal(concert_id: str):
    with get_conn() as conn:
        concert = conn.execute("SELECT * FROM concerts WHERE id = ?", (concert_id,)).fetchone()
        latest = conn.execute(
            "SELECT * FROM price_snapshots WHERE concert_id = ? ORDER BY timestamp DESC LIMIT 1", (concert_id,)
        ).fetchone()
        avg_listings_7d = conn.execute(
            "SELECT AVG(listings_count) AS avg_listings FROM price_snapshots WHERE concert_id = ?", (concert_id,)
        ).fetchone()["avg_listings"] or 0
        median_30d = conn.execute(
            "SELECT AVG(median_price) AS median_30d FROM price_snapshots WHERE concert_id = ?", (concert_id,)
        ).fetchone()["median_30d"] or 0

    if not concert or not latest:
        return jsonify({"signal": None})

    event_dt = datetime.fromisoformat(concert["event_date"].replace("Z", "+00:00"))
    days_to_concert = (event_dt - datetime.now(timezone.utc)).days

    result = generate_buy_signal(dict(latest), avg_listings_7d, median_30d, days_to_concert)
    return jsonify(result or {"signal": "HOLD", "confidence_pct": 0, "rationale": "Insufficient setup"})


@app.post("/api/export")
def export_excel():
    data = request.get_json(force=True)
    concert_id = data.get("concert_id")

    with get_conn() as conn:
        trades = [dict(r) for r in conn.execute("SELECT * FROM user_trades WHERE concert_id = ?", (concert_id,)).fetchall()]
        history = [
            dict(r)
            for r in conn.execute(
                "SELECT timestamp, get_in_price, median_price, price_change_24h_pct FROM price_snapshots WHERE concert_id = ? ORDER BY timestamp DESC",
                (concert_id,),
            ).fetchall()
        ]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Portfolio Summary"
    ws_summary["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

    stats = calculate_portfolio_stats(trades) or {}
    metric_rows = [
        ("Total Trades", stats.get("total_trades", 0)),
        ("Win Rate %", stats.get("win_rate_pct", 0)),
        ("Total Profit", stats.get("total_profit", 0)),
        ("Sharpe Ratio", stats.get("sharpe_ratio", 0)),
    ]

    for idx, (name, value) in enumerate(metric_rows, start=3):
        ws_summary[f"A{idx}"] = name
        ws_summary[f"B{idx}"] = value

    ws_trades = wb.create_sheet("Trades")
    ws_trades.append(["Concert", "Buy Date", "Buy Price", "Qty", "Sell Date", "Sell Price", "Profit", "Profit %"])
    for row in trades:
        ws_trades.append(
            [
                row.get("concert_id"),
                row.get("buy_date"),
                row.get("buy_price"),
                row.get("buy_qty"),
                row.get("sell_date"),
                row.get("sell_price"),
                row.get("profit"),
                row.get("profit_pct"),
            ]
        )

    ws_hist = wb.create_sheet("Price History")
    ws_hist.append(["Timestamp", "Get-In Price", "Median Price", "Change 24h %"])
    for row in history:
        ws_hist.append([row.get("timestamp"), row.get("get_in_price"), row.get("median_price"), row.get("price_change_24h_pct")])

    output_path = Path("/tmp") / f"ticket_tracker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_path)
    return send_file(output_path, as_attachment=True, download_name=output_path.name)


def _insert_snapshot(concert_id: str, snapshot: dict) -> None:
    with get_conn() as conn:
        previous = conn.execute(
            "SELECT get_in_price, median_price, timestamp FROM price_snapshots WHERE concert_id = ? ORDER BY timestamp DESC LIMIT 1",
            (concert_id,),
        ).fetchone()

        get_in_price = snapshot.get("get_in_price")
        median_price = snapshot.get("median_price")

        change_24h = 0.0
        change_7d = 0.0
        if previous and previous["get_in_price"]:
            change_24h = ((get_in_price - previous["get_in_price"]) / previous["get_in_price"]) * 100
        if previous and previous["median_price"]:
            change_7d = ((median_price - previous["median_price"]) / previous["median_price"]) * 100

        conn.execute(
            """
            INSERT INTO price_snapshots
            (concert_id, timestamp, source, get_in_price, median_price, listings_count, price_change_24h_pct, price_change_7d_pct)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                concert_id,
                snapshot["timestamp"],
                snapshot.get("source", "tickpick"),
                get_in_price,
                median_price,
                snapshot.get("listings_count"),
                round(change_24h, 2),
                round(change_7d, 2),
            ),
        )


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5001)
